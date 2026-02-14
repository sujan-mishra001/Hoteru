"""
Reports and export routes with branch isolation
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from datetime import datetime, timedelta
from typing import Optional

from sqlalchemy import func
from app.db.database import get_db
from app.core.dependencies import get_current_user
from app.models import Order, Product, Customer, User, Branch, Table, Floor, MenuItem, OrderItem
from app.models.purchase import PurchaseBill, PurchaseReturn, Supplier, PurchaseBillItem


from app.utils.pdf_generator import generate_pdf_report, generate_invoice_pdf, generate_multi_table_pdf_report

from app.utils.excel_generator import generate_excel_report

def get_branch_metadata(current_user, db):
    """Helper to get branch info for current session"""
    if current_user.current_branch_id:
        branch = db.query(Branch).filter(Branch.id == current_user.current_branch_id).first()
        if branch:
            return {
                "branch_name": branch.name,
                "branch_address": branch.address or branch.location,
                "branch_phone": branch.phone,
                "branch_email": branch.email
            }
    return None


def get_branch_filter(current_user):
    """Helper to get branch_id filter for queries"""
    return current_user.current_branch_id


def apply_branch_filter_order(query, branch_id):
    """Apply branch_id filter to Order queries"""
    if branch_id is not None:
        query = query.filter(Order.branch_id == branch_id)
    return query


def apply_branch_filter_table(query, branch_id):
    """Apply branch_id filter to Table queries"""
    if branch_id is not None:
        query = query.filter(Table.branch_id == branch_id)
    return query


def apply_branch_filter_floor(query, branch_id):
    """Apply branch_id filter to Floor queries"""
    if branch_id is not None:
        query = query.filter(Floor.branch_id == branch_id)
    return query


router = APIRouter()


@router.get("/dashboard-summary")
async def get_dashboard_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get summarized data for the admin dashboard with branch isolation and optional date range"""
    from datetime import datetime, time, timedelta
    
    branch_id = get_branch_filter(current_user)
    
    # Apply branch filter to tables
    table_query = db.query(Table)
    table_query = apply_branch_filter_table(table_query, branch_id)
    total_tables = table_query.count()
    occupied_tables = table_query.filter(Table.status == 'Occupied').count()
    occupancy = (occupied_tables / total_tables * 100) if total_tables > 0 else 0
    
    # Date filtering logic with branch isolation
    if start_date and end_date:
        try:
            start_dt = datetime.combine(datetime.strptime(start_date, '%Y-%m-%d').date(), time.min)
            end_dt = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), time.max)
            order_query = db.query(Order).filter(Order.created_at.between(start_dt, end_dt))
            order_query = apply_branch_filter_order(order_query, branch_id)
            orders_list = order_query.all()
            period_label = f"{start_date} to {end_date}"
            
            # For peak time data relative to the selected start date
            base_time = end_dt 
        except ValueError:
            # Fallback to 24h if date parsing fails
            base_time = datetime.now()
            start_dt = base_time - timedelta(hours=24)
            order_query = db.query(Order).filter(Order.created_at >= start_dt)
            order_query = apply_branch_filter_order(order_query, branch_id)
            orders_list = order_query.all()
            period_label = "Last 24 Hours"
    else:
        # Default to last 24 hours
        base_time = datetime.now()
        start_dt = base_time - timedelta(hours=24)
        order_query = db.query(Order).filter(Order.created_at >= start_dt)
        order_query = apply_branch_filter_order(order_query, branch_id)
        orders_list = order_query.all()
        period_label = "Last 24 Hours"
    
    # Sales breakdown
    sales_total = sum(order.net_amount or 0 for order in orders_list)
    paid_sales = sum(order.paid_amount or 0 for order in orders_list)
    credit_sales = sum(order.credit_amount or 0 for order in orders_list)
    discount = sum(order.discount or 0 for order in orders_list)
    
    # Order type breakdown
    dine_in_count = len([o for o in orders_list if o.order_type in ['Dine-In', 'Table']])
    takeaway_count = len([o for o in orders_list if o.order_type == 'Takeaway'])
    delivery_count = len([o for o in orders_list if o.order_type == 'Delivery'])

    # Outstanding revenue (all time credit, branch filtered)
    outstanding_query = db.query(Order)
    outstanding_query = apply_branch_filter_order(outstanding_query, branch_id)
    outstanding_revenue = sum(order.credit_amount or 0 for order in outstanding_query.all())

    # Top items with outstanding revenue (branch filtered)
    credit_orders = [o.id for o in orders_list if o.credit_amount > 0]
    
    if credit_orders:
        top_items_query = db.query(
            MenuItem.name,
            func.sum(OrderItem.quantity * OrderItem.price).label('total_credit')
        ).join(
            OrderItem, MenuItem.id == OrderItem.menu_item_id
        ).filter(
            OrderItem.order_id.in_(credit_orders)
        ).group_by(
            MenuItem.id, MenuItem.name
        ).order_by(
            func.sum(OrderItem.quantity * OrderItem.price).desc()
        ).limit(3).all()
        
        top_outstanding_items = [
            {"name": item.name, "amount": float(item.total_credit)}
            for item in top_items_query
        ]
    else:
        top_outstanding_items = []
    
    # Top selling items
    if orders_list:
        order_ids = [o.id for o in orders_list]
        top_selling_query = db.query(
            MenuItem.name,
            func.sum(OrderItem.quantity).label('total_quantity'),
            func.sum(OrderItem.quantity * OrderItem.price).label('total_revenue')
        ).join(
            OrderItem, MenuItem.id == OrderItem.menu_item_id
        ).filter(
            OrderItem.order_id.in_(order_ids)
        ).group_by(
            MenuItem.id, MenuItem.name
        ).order_by(
            func.sum(OrderItem.quantity * OrderItem.price).desc()
        ).limit(3).all()
        
        top_selling_items = [
            {
                "name": item.name, 
                "quantity": int(item.total_quantity),
                "revenue": float(item.total_revenue)
            }
            for item in top_selling_query
        ]
    else:
        top_selling_items = []

    # Sales by area/floor (branch filtered)
    floor_query = db.query(Floor)
    floor_query = apply_branch_filter_floor(floor_query, branch_id)
    floors = floor_query.all()
    sales_by_area = []
    for floor in floors:
        floor_tables = [t.id for t in floor.tables] if floor.tables else []
        floor_sales = sum(
            order.net_amount or 0 
            for order in orders_list 
            if order.table_id in floor_tables
        )
        if floor_sales > 0:
            sales_by_area.append({
                "area": floor.name,
                "amount": floor_sales
            })
    
    # Peak time data - 24 hourly slots
    peak_time_data = [0] * 24
    hourly_sales = [0.0] * 24
    
    for order in orders_list:
        if order.created_at:
            # Calculate hour of the day (0-23)
            hour = order.created_at.hour
            peak_time_data[hour] += 1
            hourly_sales[hour] += (order.net_amount or 0)

    return {
        "occupancy": round(occupancy, 1),
        "total_tables": total_tables,
        "occupied_tables": occupied_tables,
        "sales_24h": sales_total,
        "paid_sales": paid_sales,
        "credit_sales": credit_sales,
        "discount": discount,
        "orders_24h": len(orders_list),
        "dine_in_count": dine_in_count,
        "takeaway_count": takeaway_count,
        "delivery_count": delivery_count,
        "outstanding_revenue": outstanding_revenue,
        "top_outstanding_items": top_outstanding_items,
        "top_selling_items": top_selling_items,
        "sales_by_area": sales_by_area,
        "peak_time_data": peak_time_data,
        "hourly_sales": hourly_sales,
        "period": period_label
    }




@router.get("/sales-summary")
async def get_sales_summary(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get sales summary for the current user's branch"""
    branch_id = get_branch_filter(current_user)
    
    order_query = db.query(Order).filter(Order.status == 'Completed')
    order_query = apply_branch_filter_order(order_query, branch_id)
    orders = order_query.all()
    
    total_sales = sum(order.total_amount or 0 for order in orders)
    total_orders = len(orders)
    return {
        "total_sales": total_sales,
        "total_orders": total_orders,
        "average_order_value": total_sales / total_orders if total_orders > 0 else 0
    }


@router.get("/day-book")
async def get_day_book(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get day book (all transactions) for the current user's branch with balance calculation"""
    branch_id = get_branch_filter(current_user)
    
    query = db.query(Order)
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Order.created_at >= start_dt, Order.created_at < end_dt)
    else:
        today = datetime.now().date()
        query = query.filter(Order.created_at >= today)
        
    query = apply_branch_filter_order(query, branch_id)
    orders = query.filter(Order.status != 'Cancelled').order_by(Order.created_at.desc()).all()
    
    # Calculate running balance if needed, but for now just return list
    result = []
    total_paid = 0
    total_received = 0
    
    for o in orders:
        paid = float(o.paid_amount or 0)
        received = paid # In many cases, paid by customer is our received
        total_paid += paid
        total_received += received
        
        result.append({
            "date": o.created_at.strftime('%Y-%m-%d'),
            "paid": paid,
            "received": received,
            "balance": received - paid, # Simplified balance logic
            "order_number": o.order_number
        })
        
    return {
        "items": result,
        "summary": {
            "total_paid": total_paid,
            "total_received": total_received
        }
    }

@router.get("/daily-sales")
async def get_daily_sales_report(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get summarized daily sales report for a date range"""
    branch_id = get_branch_filter(current_user)
    
    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
    end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
    
    query = db.query(
        func.date(Order.created_at).label('date'),
        func.sum(Order.total_amount).label('gross_total'),
        func.sum(Order.discount).label('discount'),
        func.sum(Order.service_charge_amount).label('service_charge'),
        func.sum(Order.tax_amount).label('tax'),
        func.sum(Order.net_amount).label('net_total'),
        func.sum(Order.paid_amount).label('paid'),
        func.sum(Order.credit_amount).label('credit_sales'),
        func.sum(Order.delivery_charge).label('delivery_charge')
    ).filter(
        Order.created_at >= start_dt,
        Order.created_at < end_dt,
        Order.status != 'Cancelled'
    )
    
    if branch_id:
        query = query.filter(Order.branch_id == branch_id)
        
    daily_stats = query.group_by(func.date(Order.created_at)).order_by(func.date(Order.created_at).desc()).all()
    
    result = []
    total_gross = 0
    total_net = 0
    total_paid = 0
    total_credit = 0
    total_discount = 0
    
    for stat in daily_stats:
        res = {
            "date": str(stat.date),
            "gross_total": float(stat.gross_total or 0),
            "discount": float(stat.discount or 0),
            "complementary": 0, # Placeholder if not tracked specifically
            "delivery_commission": 0, # Placeholder
            "net_total": float(stat.net_total or 0),
            "paid": float(stat.paid or 0),
            "credit_sales": float(stat.credit_sales or 0),
            "net_delivery": float(stat.delivery_charge or 0),
            "credit_service": 0, # Placeholder
            "cash": 0, # Need payment type breakdown for these
            "fonepay": 0,
            "esewa": 0
        }
        
        # Payment breakdown for this specific day
        day_start = datetime.combine(stat.date, datetime.min.time())
        day_end = datetime.combine(stat.date, datetime.max.time())
        
        breakdown = db.query(
            Order.payment_type,
            func.sum(Order.paid_amount).label('amount')
        ).filter(
            Order.created_at.between(day_start, day_end),
            Order.status != 'Cancelled'
        )
        if branch_id:
            breakdown = breakdown.filter(Order.branch_id == branch_id)
            
        breakdown = breakdown.group_by(Order.payment_type).all()
        
        for b in breakdown:
            ptype = (b.payment_type or "").lower()
            if 'cash' in ptype: res['cash'] += float(b.amount or 0)
            elif 'fonepay' in ptype: res['fonepay'] += float(b.amount or 0)
            elif 'esewa' in ptype: res['esewa'] += float(b.amount or 0)

        total_gross += res['gross_total']
        total_net += res['net_total']
        total_paid += res['paid']
        total_credit += res['credit_sales']
        total_discount += res['discount']
        
        result.append(res)
        
    return {
        "items": result,
        "summary": {
            "gross_sales": total_gross,
            "discount": total_discount,
            "net_sales": total_net,
            "paid_sales": total_paid,
            "credit_sales": total_credit,
            "complementary": 0,
            "delivery_commission": 0,
            "net_delivery": sum(item['net_delivery'] for item in result)
        }
    }

@router.get("/monthly-sales")
async def get_monthly_sales_report(
    year: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get summarized monthly sales report for a year"""
    branch_id = get_branch_filter(current_user)
    
    # Months list for display
    months = ["Baisakh", "Jestha", "Ashad", "Shrawan", "Bhadra", "Ashwin", "Kartik", "Mangsir", "Poush", "Magh", "Falgun", "Chaitra"]
    
    # Since we are using standard datetime, we'll map Gregorian months to a year view
    # For a true B.S report, a conversion library would be needed. 
    # For now, we aggregate by Gregorian months 1-12.
    
    result = {}
    
    query = db.query(
        func.extract('month', Order.created_at).label('month'),
        func.sum(Order.total_amount).label('gross_sales'),
        func.sum(Order.discount).label('discount'),
        func.sum(Order.net_amount).label('net_sales'),
        func.sum(Order.paid_amount).label('paid_sales'),
        func.sum(Order.credit_amount).label('credit_sales')
    ).filter(
        func.extract('year', Order.created_at) == year,
        Order.status != 'Cancelled'
    )
    
    if branch_id:
        query = query.filter(Order.branch_id == branch_id)
        
    monthly_stats = query.group_by(func.extract('month', Order.created_at)).all()
    
    # Prepare base structure
    monthly_map = {int(stat.month): stat for stat in monthly_stats}
    
    rows = [
        "Gross Sales", "Discount", "Complementary", "Net Sales", "Paid Sales", "Cash", "Esewa", "Fonepay", "Customer Credit"
    ]
    
    # For simplicity, we'll return a structure that the frontend can easily map to a table
    # Columns: Months (1-12), Rows: Metrics
    
    final_data = []
    for row_name in rows:
        row_obj = {"particular": row_name}
        for m_idx in range(1, 13):
            month_label = f"month_{m_idx}"
            stat = monthly_map.get(m_idx)
            
            val = 0
            if stat:
                if row_name == "Gross Sales": val = float(stat.gross_sales or 0)
                elif row_name == "Discount": val = float(stat.discount or 0)
                elif row_name == "Net Sales": val = float(stat.net_sales or 0)
                elif row_name == "Paid Sales": val = float(stat.paid_sales or 0)
                elif row_name == "Customer Credit": val = float(stat.credit_sales or 0)
                elif row_name in ["Cash", "Esewa", "Fonepay"]:
                    # Payment type specific monthly aggregation
                    p_query = db.query(func.sum(Order.paid_amount)).filter(
                        func.extract('year', Order.created_at) == year,
                        func.extract('month', Order.created_at) == m_idx,
                        Order.payment_type.ilike(f"%{row_name}%"),
                        Order.status != 'Cancelled'
                    )
                    if branch_id: p_query = p_query.filter(Order.branch_id == branch_id)
                    val = float(p_query.scalar() or 0)
            
            row_obj[month_label] = val
        final_data.append(row_obj)
        
    return final_data

@router.get("/purchase-report")
async def get_purchase_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    supplier_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get summarized purchase report"""
    branch_id = get_branch_filter(current_user)
    
    query = db.query(PurchaseBill).options(joinedload(PurchaseBill.supplier))
    
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(PurchaseBill.order_date >= start_dt, PurchaseBill.order_date < end_dt)
        
    if supplier_id:
        query = query.filter(PurchaseBill.supplier_id == supplier_id)
        
    if branch_id:
        query = query.filter(PurchaseBill.branch_id == branch_id)
        
    bills = query.order_by(PurchaseBill.order_date.desc()).all()
    
    result = []
    total_payable = 0
    total_paid = 0
    
    for b in bills:
        # Simplified assumption: If status is 'Paid', entire amount is paid. 
        # In a more complex system, we'd have a 'paid_amount' field on PurchaseBill.
        paid = float(b.total_amount) if getattr(b, 'status', '').lower() == 'paid' else 0
        total_payable += float(b.total_amount)
        total_paid += paid
        
        result.append({
            "bill_number": b.bill_number,
            "date": b.order_date.strftime('%Y-%m-%d'),
            "supplier_name": b.supplier.name if b.supplier else "N/A",
            "payable": float(b.total_amount),
            "paid": paid,
            "status": b.status or "Pending",
            "paid_by": "System" # Placeholder
        })
        
    return {
        "items": result,
        "summary": {
            "total_payable": total_payable,
            "total_paid": total_paid,
            "total_bills": len(bills)
        }
    }


@router.get("/export/pdf/{report_type}")
async def export_pdf(
    report_type: str,
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Export report as PDF with optional date filtering"""
    from datetime import datetime, time as dtime
    
    # Helper for date filtering
    query_start = None
    query_end = None
    
    if date:
        dt = datetime.strptime(date, '%Y-%m-%d')
        query_start = datetime.combine(dt.date(), dtime.min)
        query_end = datetime.combine(dt.date(), dtime.max)
    elif start_date and end_date:
        query_start = datetime.combine(datetime.strptime(start_date, '%Y-%m-%d').date(), dtime.min)
        query_end = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), dtime.max)
    if report_type == "session":
        return await export_sessions_pdf(db, current_user)
    elif report_type == "user":
        report_type = "staff"

    if report_type == "sales-summary":
        result = await get_sales_summary(db, current_user)
        data = [{"Metric": k, "Value": v} for k, v in result.items()]
        title = "Sales Summary"
    elif report_type == "day-book":
        orders = await get_day_book(db, current_user)
        data = [{"Order Number": o.order_number, "Total": o.total_amount, "Date": str(o.created_at)} for o in orders]
        title = "Day Book"
    elif report_type == "sales":
        # Get orders based on date selection
        if query_start and query_end:
            title = f"Sales Report ({query_start.strftime('%Y-%m-%d')})" if date else f"Sales Report ({start_date} to {end_date})"
            orders = db.query(Order).options(joinedload(Order.items).joinedload(OrderItem.menu_item)).filter(Order.created_at.between(query_start, query_end)).all()
        else:
            # Default to last 24 hours
            last_24h = datetime.now() - timedelta(hours=24)
            orders = db.query(Order).options(joinedload(Order.items).joinedload(OrderItem.menu_item)).filter(Order.created_at >= last_24h).all()
            title = "Daily Sales Report (Last 24h)"
        
        # Summary Metrics
        summary_metrics = {
            "Total Orders": len(orders),
            "Gross Sales (Items)": sum(o.gross_amount or 0 for o in orders),
            "Total Discount": sum(o.discount or 0 for o in orders),
            "Service Charge": sum(o.service_charge_amount or 0 for o in orders),
            "VAT (Tax)": sum(o.tax_amount or 0 for o in orders),
            "Net Sales (Total Payable)": sum(o.net_amount or 0 for o in orders),
            "Paid Amount": sum(o.paid_amount or 0 for o in orders),
            "Credit (Outstanding)": sum(o.credit_amount or 0 for o in orders),
        }
        
        summary_data = [{"Metric": k, "Value": f"Rs. {v:,.2f}" if isinstance(v, (int, float)) and k != "Total Orders" else v} for k, v in summary_metrics.items()]
        
        # Add payment type breakdown
        payment_breakdown = {}
        for o in orders:
            if o.paid_amount > 0:
                pt = o.payment_type or "Cash"
                payment_breakdown[pt] = payment_breakdown.get(pt, 0) + o.paid_amount
        
        # Item Performance breakdown
        item_stats = {}
        for o in orders:
            for item in o.items:
                name = item.menu_item.name if item.menu_item else "Unknown"
                if name not in item_stats:
                    item_stats[name] = {"qty": 0, "rev": 0}
                item_stats[name]["qty"] += item.quantity
                item_stats[name]["rev"] += (item.quantity * item.price)
        
        item_data = []
        for name, stats in sorted(item_stats.items(), key=lambda x: x[1]['rev'], reverse=True):
            item_data.append({
                "Item Name": name,
                "Qty Sold": stats['qty'],
                "Revenue": f"Rs. {stats['rev']:,.2f}"
            })

        sections = [
            {"title": "Financial Summary", "data": summary_data},
            {"title": "Top Selling Items (Performance)", "data": item_data}
        ]
        
        # Calculate Period String properly
        period_str = f"{query_start.strftime('%Y-%m-%d')} to {query_end.strftime('%Y-%m-%d')}" if query_start and query_end else "Last 24 Hours"
        if date and query_start:
             period_str = query_start.strftime('%Y-%m-%d')

        metadata = get_branch_metadata(current_user, db) or {}
        metadata['period'] = period_str

        pdf_buffer = generate_multi_table_pdf_report(sections, title=title.split('(')[0].strip(), metadata=metadata)
        return StreamingResponse(pdf_buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=sales_detailed.pdf"})

    elif report_type == "inventory":
        from app.models import InventoryTransaction
        products = db.query(Product).all()
        
        # Calculate stock movements (basic aggregation for all time)
        # Note: Ideally this would be optimized with SQL queries, but reusing python logic for consistency
        products_data = []
        for p in products:
            added = 0.0
            consumed = 0.0
            
            # This is expensive (N+1), but simple given current ORM loading. 
            # Given user wants "Added" and "Available", we iterate transactions.
            for txn in p.transactions:
                if txn.transaction_type in ['IN', 'Add', 'Production_IN', 'Adjustment']:
                     # Assuming Adjustment is positive here, or we check sign. 
                     # Product.current_stock treats Adjustment as add/sub depending on sign? No, it adds it.
                     if txn.quantity > 0:
                        added += txn.quantity
                elif txn.transaction_type in ['OUT', 'Remove', 'Production_OUT']:
                    consumed += txn.quantity
            
            products_data.append({
                "Product": p.name, 
                "Category": p.category or "-", 
                "Added Stock": round(added, 3),
                "Used/Sold": round(consumed, 3),
                "Available": round(p.current_stock, 3),
                "Unit": p.unit.abbreviation if p.unit else "-",
            })
        data = products_data
        title = "Inventory Stock & Consumption Report"
    elif report_type == "customers":
        customers = db.query(Customer).all()
        data = [{
            "Customer": c.name, 
            "Phone": c.phone or "-", 
            "Type": c.customer_type or "Regular",
            "Visits": c.total_visits or 0,
            "Total Spent": f"Rs. {c.total_spent:,.2f}" if c.total_spent else "Rs. 0.00",
            "Due": f"Rs. {c.due_amount:,.2f}" if c.due_amount else "Rs. 0.00"
        } for c in customers]
        title = "Customer Analysis & Loyalty"
    elif report_type == "staff":
        users = db.query(User).all()
        data = [{"Staff": u.full_name, "Role": u.role, "Username": u.username, "Status": "Active" if not u.disabled else "Disabled"} for u in users]
        title = "Staff Account List"
    elif report_type == "purchase":
        branch_id = get_branch_filter(current_user)
        query = db.query(PurchaseBill).options(joinedload(PurchaseBill.items).joinedload(PurchaseBillItem.product))
        if branch_id:
            query = query.filter(PurchaseBill.branch_id == branch_id)
        
        bills = query.filter(PurchaseBill.order_date.between(query_start, query_end)).all() if query_start and query_end else query.all()
        
        bill_data = []
        item_data = []
        for b in bills:
            bill_data.append({
                "Bill #": b.bill_number,
                "Supplier": b.supplier.name if b.supplier else "N/A",
                "Date": b.order_date.strftime('%Y-%m-%d'),
                "Amount": f"Rs. {b.total_amount:,.2f}"
            })
            for item in b.items:
                item_data.append({
                    "Date": b.order_date.strftime('%Y-%m-%d'),
                    "Product": item.product.name if item.product else "Unknown",
                    "Qty": item.quantity,
                    "Rate": f"Rs. {item.rate:,.2f}",
                    "Subtotal": f"Rs. {(item.quantity * item.rate):,.2f}",
                    "Bill #": b.bill_number
                })
        
        sections = [
            {"title": "Purchase Bills Summary", "data": bill_data},
            {"title": "Detailed Purchase Items (Materials)", "data": item_data}
        ]
        pdf_buffer = generate_multi_table_pdf_report(sections, title="Purchase Detailed Report", metadata=get_branch_metadata(current_user, db))
        return StreamingResponse(pdf_buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=purchase_detailed.pdf"})
    else:


        raise HTTPException(status_code=404, detail="Report type not found")
    
    metadata = get_branch_metadata(current_user, db) or {}
    metadata['period'] = title.split('(')[-1].replace(')', '') if '(' in title else "Last 24 Hours"
    
    pdf_buffer = generate_pdf_report(data, title.split('(')[0].strip(), metadata=metadata)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={report_type}.pdf"}
    )


@router.get("/export/excel/{report_type}")
async def export_excel(
    report_type: str,
    date: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Export report as Excel with optional date filtering"""
    from datetime import datetime, time as dtime
    from app.models import User, MenuItem, Order, Customer # Keep these as they are used
    from app.models.inventory import InventoryTransaction, BatchProduction, BillOfMaterials
    from app.models.pos_session import POSSession
    
    query_start = None
    query_end = None
    
    if date:
        dt = datetime.strptime(date, '%Y-%m-%d')
        query_start = datetime.combine(dt.date(), dtime.min)
        query_end = datetime.combine(dt.date(), dtime.max)
    elif start_date and end_date:
        query_start = datetime.combine(datetime.strptime(start_date, '%Y-%m-%d').date(), dtime.min)
        query_end = datetime.combine(datetime.strptime(end_date, '%Y-%m-%d').date(), dtime.max)
    metadata = get_branch_metadata(current_user, db) or {}
    metadata['period'] = "Full Summary"
    
    if report_type == "sales-summary":
        result = await get_sales_summary(db, current_user)
        data = [{"Metric": k, "Value": v} for k, v in result.items()]
        excel_buffer = generate_excel_report(data, "Sales Summary", metadata=metadata)
    elif report_type == "day-book":
        orders = await get_day_book(db, current_user)
        data = [{"Order Number": o.order_number, "Total": o.total_amount, "Date": str(o.created_at)} for o in orders]
        excel_buffer = generate_excel_report(data, "Day Book", metadata=metadata)
    elif report_type == "sales":
        if query_start and query_end:
            orders = db.query(Order).filter(Order.created_at.between(query_start, query_end)).all()
            title = f"Sales Report ({query_start.strftime('%Y-%m-%d')})" if date else f"Sales Report ({start_date} to {end_date})"
        else:
            last_24h = datetime.now() - timedelta(hours=24)
            orders = db.query(Order).filter(Order.created_at >= last_24h).all()
            title = "Daily Sales Report (Last 24h)"
        
        data = []
        for o in orders:
            data.append({
                "Order #": o.order_number,
                "Type": o.order_type,
                "Status": o.status,
                "Gross": o.gross_amount,
                "Discount": o.discount,
                "SC": o.service_charge_amount,
                "VAT": o.tax_amount,
                "Net": o.net_amount,
                "Paid": o.paid_amount,
                "Credit": o.credit_amount,
                "Payment": o.payment_type or "-",
                "Items": ", ".join([f"{i.menu_item.name} x{i.quantity}" for i in o.items if i.menu_item]),
                "Date": o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else "-"
            })
        
        if not data:
            data = [{"Message": "No sales record found for today"}]
            
        metadata = get_branch_metadata(current_user, db) or {}
        metadata['period'] = title.split('(')[-1].replace(')', '') if '(' in title else "Today"
        
        excel_buffer = generate_excel_report(data, title.split('(')[0].strip(), metadata=metadata)
    elif report_type == "inventory":
        # 1. Fetch Products with Unit
        products = db.query(Product).options(joinedload(Product.unit)).all()
        products_map = {p.id: p for p in products}
        
        # 2. Fetch Transactions with User and batch info for consumption tracking
        
        # We need to map reference_id to BatchProduction to find out what was produced
        # This is optimization heavy. Let's fetch all relevant BatchProductions first.
        # But reference_id in InventoryTransaction is generic.
        # Assuming reference_number starts with "PROD-" for production.
        
        txns_query = db.query(InventoryTransaction).options(
            joinedload(InventoryTransaction.user)
        ).order_by(InventoryTransaction.created_at.asc())
        
        all_txns = txns_query.all()
        
        # Prefetch BatchProductions to map to Menu Items
        prod_txns = [t.reference_id for t in all_txns if t.transaction_type == 'Production_OUT' and t.reference_id]
        batch_map = {}
        if prod_txns:
            batches = db.query(BatchProduction).options(
                joinedload(BatchProduction.bom).joinedload(BillOfMaterials.menu_items),
                joinedload(BatchProduction.finished_product)
            ).filter(BatchProduction.id.in_(prod_txns)).all()
            batch_map = {b.id: b for b in batches}
        
        # 3. Process Data
        summary_stats = {p.id: {
            "opening": 0.0,
            "added": 0.0,
            "produced": 0.0,
            "consumed": 0.0,
            "sold": 0.0,
            "adjusted": 0.0,
            "last_txn_date": "-",
            "consumption_breakdown": {} # Map of "Menu Item Name" -> Qty Used
        } for p in products}

        detailed_log = []

        query_start_dt = None
        query_end_dt = None
        if query_start and query_end:
            query_start_dt = query_start
            query_end_dt = query_end
        
        for txn in all_txns:
            pid = txn.product_id
            if pid not in summary_stats: continue
            
            qty = float(txn.quantity or 0)
            t_type = txn.transaction_type
            t_date = txn.created_at
            
            # Opening Stock Logic
            if query_start_dt and t_date < query_start_dt:
                if t_type in ['IN', 'Add', 'Production_IN', 'Adjustment', 'Count']: 
                     summary_stats[pid]["opening"] += qty
                elif t_type in ['OUT', 'Remove', 'Production_OUT']:
                     summary_stats[pid]["opening"] -= qty
                continue
            
            # Future Transaction Check
            if query_end_dt and t_date > query_end_dt:
                continue
            
            # Categorization
            if t_type in ['IN', 'Add']:
                summary_stats[pid]["added"] += qty
            elif t_type == 'Production_IN':
                summary_stats[pid]["produced"] += qty
            elif t_type == 'Production_OUT':
                summary_stats[pid]["consumed"] += qty
            elif t_type in ['OUT', 'Remove']:
                summary_stats[pid]["sold"] += qty
            elif t_type in ['Adjustment', 'Count']:
                summary_stats[pid]["adjusted"] += qty
            
            # Track Usage Breakdown for Production_OUT
            if t_type == 'Production_OUT' and txn.reference_id in batch_map:
                batch = batch_map[txn.reference_id]
                # Identify produced item name
                produced_name = "Unknown"
                if batch.bom and batch.bom.menu_items:
                    produced_name = ", ".join([mi.name for mi in batch.bom.menu_items])
                elif batch.finished_product:
                    produced_name = batch.finished_product.name
                elif batch.bom:
                    produced_name = batch.bom.name
                
                if produced_name:
                    if produced_name not in summary_stats[pid]["consumption_breakdown"]:
                        summary_stats[pid]["consumption_breakdown"][produced_name] = 0.0
                    summary_stats[pid]["consumption_breakdown"][produced_name] += qty

            summary_stats[pid]["last_txn_date"] = t_date.strftime('%Y-%m-%d')
            
            # Detailed Log
            p = products_map[pid]
            detailed_log.append({
                "Date": t_date.strftime('%Y-%m-%d %H:%M'),
                "Product": p.name,
                "Type": t_type,
                "Quantity": qty,
                "Unit": p.unit.abbreviation if p.unit else "-",
                "User": txn.user.full_name if txn.user else "System",
                "Ref #": txn.reference_number or "-",
                "Notes": txn.notes or "-"
            })

        # Build aggregated menu item production stats
        from app.models import OrderItem, Order
        
        # Fetch all productions with their BOMs and menu items
        all_batches = db.query(BatchProduction).options(
            joinedload(BatchProduction.bom).joinedload(BillOfMaterials.menu_items),
            joinedload(BatchProduction.finished_product)
        ).all()
        
        # Aggregate by menu item
        menu_item_stats = {}  # {menu_item_id: {name, produced, sold, remaining}}
        
        for batch in all_batches:
            if not batch.bom:
                continue
                
            total_produced = batch.bom.output_quantity * batch.quantity
            
            # Get menu items for this batch
            menu_items = batch.bom.menu_items if batch.bom.menu_items else []
            if not menu_items and batch.finished_product:
                # If no menu items, check if finished_product itself is tracked
                # For now, skip if no menu items
                continue
            
            for mi in menu_items:
                if mi.id not in menu_item_stats:
                    # Calculate total sold for this menu item
                    sold = db.query(func.sum(OrderItem.quantity)).join(Order).filter(
                        OrderItem.menu_item_id == mi.id,
                        Order.status != 'Cancelled'
                    ).scalar()
                    total_sold = float(sold) if sold else 0.0
                    
                    # Calculate total produced for this menu item across all batches
                    total_prod_for_item = 0.0
                    for b in all_batches:
                        if b.bom and any(m.id == mi.id for m in b.bom.menu_items):
                            total_prod_for_item += b.bom.output_quantity * b.quantity
                    
                    menu_item_stats[mi.id] = {
                        "name": mi.name,
                        "produced": total_prod_for_item,
                        "sold": total_sold,
                        "remaining": total_prod_for_item - total_sold
                    }
        
        data = []
        for p in products:
            stats = summary_stats[p.id]
            movements = (stats["added"] + stats["produced"] + stats["adjusted"]) - (stats["consumed"] + stats["sold"])
            closing = stats["opening"] + movements
            
            # Determine which menu item consumed this raw material the most
            produced_item = "-"
            produced_qty = 0.0
            sold_qty = 0.0
            remaining_qty = 0.0
            
            if stats["consumption_breakdown"]:
                # Get the item with highest consumption
                top_item_name = max(stats["consumption_breakdown"].items(), key=lambda x: x[1])[0]
                
                # Find this menu item in our stats
                for mi_id, mi_stats in menu_item_stats.items():
                    if mi_stats["name"] == top_item_name or top_item_name in mi_stats["name"]:
                        produced_item = mi_stats["name"]
                        produced_qty = mi_stats["produced"]
                        sold_qty = mi_stats["sold"]
                        remaining_qty = mi_stats["remaining"]
                        break
            
            data.append({
                "Product": p.name,
                "Category": p.category or "-",
                "Unit": p.unit.abbreviation if p.unit else "-",
                "Added": round(stats["added"], 2),
                "Consumed": round(stats["consumed"], 2),
                "Adjusted": round(stats["adjusted"], 2),
                "Closing": round(closing, 2),
                "Last Txn": stats["last_txn_date"]
            })
        
        # Build Item Tracking data (one row per menu item)
        item_tracking_data = []
        for mi_id, mi_stats in menu_item_stats.items():
            item_tracking_data.append({
                "Menu Item": mi_stats["name"],
                "Produced Quantity": round(mi_stats["produced"], 2),
                "Sold Quantity": round(mi_stats["sold"], 2),
                "Remaining Quantity": round(mi_stats["remaining"], 2)
            })
            
        import pandas as pd
        from io import BytesIO
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill
        
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Sheet 1: Inventory Tracking
            df_inventory = pd.DataFrame(data)
            df_inventory.to_excel(writer, sheet_name='Inventory Tracking', index=False, startrow=1)
            
            worksheet_inv = writer.sheets['Inventory Tracking']
            
            # Add header
            worksheet_inv.merge_cells('A1:H1')
            worksheet_inv['A1'] = 'INVENTORY TRACKING'
            worksheet_inv['A1'].font = Font(bold=True, size=12)
            worksheet_inv['A1'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet_inv['A1'].fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
            
            # Style column headers (row 2)
            for col_num in range(1, len(df_inventory.columns) + 1):
                cell = worksheet_inv.cell(row=2, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
            
            # Auto-adjust column widths
            for column in worksheet_inv.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet_inv.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 2: Item Tracking
            if item_tracking_data:
                df_items = pd.DataFrame(item_tracking_data)
                df_items.to_excel(writer, sheet_name='Item Tracking', index=False, startrow=1)
                
                worksheet_items = writer.sheets['Item Tracking']
                
                # Add header
                worksheet_items.merge_cells('A1:D1')
                worksheet_items['A1'] = 'ITEM TRACKING'
                worksheet_items['A1'].font = Font(bold=True, size=12)
                worksheet_items['A1'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet_items['A1'].fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
                
                # Style column headers (row 2)
                for col_num in range(1, len(df_items.columns) + 1):
                    cell = worksheet_items.cell(row=2, column=col_num)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
                
                # Auto-adjust column widths
                for column in worksheet_items.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet_items.column_dimensions[column_letter].width = adjusted_width
            
            # Sheet 3: Transaction Log
            if detailed_log:
                pd.DataFrame(detailed_log).to_excel(writer, sheet_name='Transaction Log', index=False)
            else:
                pd.DataFrame([{"Message": "No transactions found"}]).to_excel(writer, sheet_name='Transaction Log', index=False)
        
        buffer.seek(0)
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=inventory_detailed_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
    elif report_type == "customers":
        customers = db.query(Customer).all()
        data = [{
            "Name": c.name,
            "Phone": c.phone or "-",
            "Email": c.email or "-",
            "Type": c.customer_type or "Regular",
            "Visits": c.total_visits or 0,
            "Total Spent": c.total_spent or 0,
            "Due Amount": c.due_amount or 0,
            "Created At": c.created_at.strftime('%Y-%m-%d') if c.created_at else "-"
        } for c in customers]
        excel_buffer = generate_excel_report(data, "Customer Analytics", metadata=metadata)
    elif report_type == "staff":
        users = db.query(User).all()
        data = [{
            "Name": u.full_name,
            "Username": u.username,
            "Email": u.email,
            "Role": u.role,
            "Organization": u.company_name or "-",
            "Status": "Active" if not u.disabled else "Disabled"
        } for u in users]
        excel_buffer = generate_excel_report(data, "Staff", metadata=metadata)
    elif report_type == "purchase":
        branch_id = get_branch_filter(current_user)
        query = db.query(PurchaseBill)
        if branch_id:
            query = query.filter(PurchaseBill.branch_id == branch_id)
            
        if query_start and query_end:
            bills = query.filter(PurchaseBill.order_date.between(query_start, query_end)).all()
        else:
            bills = query.all()
            
        data = []
        for b in bills:
            data.append({
                "Bill #": b.bill_number,
                "Supplier": b.supplier.name if b.supplier else "-",
                "Date": b.order_date.strftime('%Y-%m-%d') if b.order_date else "-",
                "Status": b.status,
                "Branch": b.branch_id,
                "Total Amount": b.total_amount
            })
            for item in b.items:
                data.append({
                    "Bill #": f" -> Item: {item.product.name if item.product else '-'}",
                    "Supplier": "Component",
                    "Date": b.order_date.strftime('%Y-%m-%d'),
                    "Status": "Included",
                    "Total Amount": item.total_amount,
                    "Branch": f"Qty: {item.quantity} @ {item.rate}"
                })
            
        if not data:
            data = [{"Message": "No purchase records found"}]
            
        excel_buffer = generate_excel_report(data, "Purchase Detailed", metadata=metadata)

    elif report_type in ["session", "sessions"]:
        sessions = db.query(POSSession).all()
        data = []
        for s in sessions:
            data.append({
                "ID": s.id,
                "Staff": s.user.full_name if s.user else "System",
                "Start": s.start_time.strftime('%Y-%m-%d %H:%M') if s.start_time else "-",
                "End": s.end_time.strftime('%Y-%m-%d %H:%M') if s.end_time else "Open",
                "Orders": s.total_orders or 0,
                "Gross Sales": s.total_sales or 0,
                "Net Total": s.net_total or 0,
                "Opening Cash": s.opening_cash or 0,
                "Expected Cash": s.expected_cash or 0,
                "Actual Cash": s.actual_cash or 0,
                "Difference": (s.actual_cash or 0) - (s.expected_cash or 0),
                "Status": s.status
            })
        excel_buffer = generate_excel_report(data, "POS Session Detail Analysis", metadata=metadata)

    else:
        raise HTTPException(status_code=404, detail="Report type not found")
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={report_type}.xlsx"}
    )


@router.get("/orders/{order_id}/invoice")
async def get_order_invoice(
    order_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Generate invoice PDF for an order"""
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order_data = {
        "order_number": order.order_number,
        "customer": {"name": order.customer.name if order.customer else None},
        "table": {"table_id": order.table.table_id if order.table else None},
        "items": [{"name": item.menu_item.name, "quantity": item.quantity, "price": item.price, "subtotal": item.subtotal} for item in order.items]
    }
    
    metadata = get_branch_metadata(current_user, db)
    pdf_buffer = generate_invoice_pdf(order_data, metadata=metadata)
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=invoice_{order.order_number}.pdf"}
    )


@router.get("/export/all/excel")
async def export_all_excel(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Export all report data into a single multi-sheet Excel file"""
    import pandas as pd
    from io import BytesIO
    from datetime import datetime, timedelta
    
    buffer = BytesIO()
    last_24h = datetime.now() - timedelta(hours=24)
    
    # 1. Sales Data
    orders = db.query(Order).filter(Order.created_at >= last_24h).all()
    sales_df = pd.DataFrame([{
        "Order #": o.order_number,
        "Type": o.order_type,
        "Status": o.status,
        "Net Amount": o.net_amount,
        "Payment": o.payment_type or "Cash",
        "Time": o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else "-"
    } for o in orders])
    
    # 2. Inventory Data
    products = db.query(Product).all()
    inventory_df = pd.DataFrame([{
        "Product": p.name,
        "Category": p.category or "-",
        "Stock": p.current_stock,
        "Min Stock": p.min_stock,
        "Status": p.status
    } for p in products])
    
    # 3. Customer Data
    customers = db.query(Customer).all()
    customers_df = pd.DataFrame([{
        "Name": c.name,
        "Phone": c.phone or "-",
        "Type": c.customer_type or "Regular",
        "Visits": c.total_visits or 0,
        "Total Spent": c.total_spent or 0,
        "Due": c.due_amount or 0
    } for c in customers])
    
    # 4. Staff Data
    users = db.query(User).all()
    staff_df = pd.DataFrame([{
        "Name": u.full_name,
        "Username": u.username,
        "Role": u.role,
        "Status": "Active" if not u.disabled else "Disabled"
    } for u in users])
    
    # 5. Purchase Data
    purchases = db.query(PurchaseBill).filter(PurchaseBill.order_date >= last_24h).all()
    purchase_df = pd.DataFrame([{
        "Bill #": p.bill_number,
        "Supplier": p.supplier.name if p.supplier else "-",
        "Date": p.order_date.strftime('%Y-%m-%d') if p.order_date else "-",
        "Status": p.status,
        "Total": p.total_amount
    } for p in purchases])
    
    metadata = get_branch_metadata(current_user, db)

    header_rows = 0
    header_df = pd.DataFrame()
    
    if metadata:
        header_data = [
            [metadata.get('branch_name', '').upper()],
            [metadata.get('branch_address', '')],
            [f"Tel: {metadata.get('branch_phone', '')} | Email: {metadata.get('branch_email', '')}"],
            ["MASTER BUSINESS REPORT"],
            [f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            [""] # Spacer
        ]
        header_df = pd.DataFrame(header_data)
        header_rows = len(header_data)

    # Ensure at least one sheet exists
    sheets = {
        'Sales Transactions': sales_df,
        'Inventory Levels': inventory_df,
        'Customer Database': customers_df,
        'Staff List': staff_df,
        'Purchase History': purchase_df
    }

    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        has_data = False
        for sheet_name, df in sheets.items():
            if not df.empty:
                has_data = True
                if not header_df.empty:
                    header_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=header_rows)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        if not has_data:
            pd.DataFrame([{"Message": "No data available"}]).to_excel(writer, sheet_name='No Data', index=False)
    
    buffer.seek(0)
    
    metadata = get_branch_metadata(current_user, db)
    prefix = metadata.get('branch_name').replace(' ', '_') if metadata and metadata.get('branch_name') else "Business"
    filename = f"{prefix}_Master_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/sessions")
def get_sessions_report(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Get all POS sessions for reporting"""
    from app.models.pos_session import POSSession
    from app.models.auth import User
    
    sessions = db.query(POSSession).order_by(POSSession.start_time.desc()).all()
    
    result = []
    for session in sessions:
        user = session.user
        result.append({
            "id": session.id,
            "user_id": session.user_id,
            "user": {
                "full_name": user.full_name if user else "Unknown",
                "role": user.role if user else "Staff"
            },
            "start_time": session.start_time.isoformat() if session.start_time else None,
            "end_time": session.end_time.isoformat() if session.end_time else None,
            "status": session.status,
            "opening_cash": session.opening_cash,
            "actual_cash": session.actual_cash,
            "expected_cash": session.expected_cash,
            "total_sales": session.total_sales,
            "total_orders": session.total_orders,
            "notes": session.notes
        })
    
    return result


@router.get("/export/sessions/pdf")
async def export_sessions_pdf(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Export all sessions as a PDF report"""
    from app.models.pos_session import POSSession
    
    sessions = db.query(POSSession).order_by(POSSession.start_time.desc()).all()
    
    data = []
    for s in sessions:
        user_name = s.user.full_name if s.user else "Unknown"
        duration = "-"
        if s.end_time and s.start_time:
            diff = s.end_time - s.start_time
            hours = diff.total_seconds() // 3600
            minutes = (diff.total_seconds() % 3600) // 60
            duration = f"{int(hours)}h {int(minutes)}m"
        elif s.status == "Open":
            duration = "Ongoing"

        data.append({
            "ID": f"#{s.id}",
            "Staff": user_name,
            "Start": s.start_time.strftime('%b %d, %H:%M') if s.start_time else "-",
            "End": s.end_time.strftime('%b %d, %H:%M') if s.end_time else "-",
            "Status": s.status,
            "Opening": f"Rs. {s.opening_cash:,.2f}",
            "Actual": f"Rs. {s.actual_cash:,.2f}",
            "Sales": f"Rs. {s.total_sales:,.2f}",
            "Orders": s.total_orders,
            "Duration": duration
        })
    
    metadata = get_branch_metadata(current_user, db)
    pdf_buffer = generate_pdf_report(data, "POS Shift Statistics Report", metadata=metadata)
    
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=session_report.pdf"}
    )


@router.get("/export/shift/{session_id}")
async def export_shift_report(
    session_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """Generate a detailed single-shift report PDF"""
    from app.models.pos_session import POSSession
    from app.models import Order
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    session = db.query(POSSession).filter(POSSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
        
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title & Header
    metadata = get_branch_metadata(current_user, db)
    
    branch_name = "BUSINESS"
    if metadata:
        branch_name = metadata.get('branch_name') or "BUSINESS"
        elements.append(Paragraph(f"<b>{branch_name.upper()}</b>", styles['Normal']))
        if metadata.get('branch_address'):
            elements.append(Paragraph(metadata.get('branch_address'), styles['Normal']))
    else:
         elements.append(Paragraph(f"<b>{branch_name}</b>", styles['Normal']))

    elements.append(Paragraph(f"<b>{branch_name} Shift Report #{session.id}</b>", styles['Title']))
    elements.append(Paragraph(f"Staff: {session.user.full_name if session.user else 'Unknown'}", styles['Normal']))
    elements.append(Paragraph(f"Status: {session.status}", styles['Normal']))
    elements.append(Paragraph(f"Start: {session.start_time.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    if session.end_time:
        elements.append(Paragraph(f"End: {session.end_time.strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    
    elements.append(Spacer(1, 20))
    
    # Financial Table
    data = [
        ["DESCRIPTION", "AMOUNT"],
        ["Opening Cash", f"Rs. {session.opening_cash:,.2f}"],
        ["Cash Sales", f"Rs. {session.cash_sales:,.2f}"],
        ["Online Sales", f"Rs. {session.online_sales:,.2f}"],
        ["Credit Sales", f"Rs. {session.credit_sales:,.2f}"],
        ["---", "---"],
        ["Expected Cash in Drawer", f"Rs. {session.expected_cash:,.2f}"],
        ["Actual Cash Reported", f"Rs. {session.actual_cash:,.2f}"],
        ["Difference", f"Rs. {(session.actual_cash - session.expected_cash):,.2f}"]
    ]
    
    t = Table(data, colWidths=[3*inch, 2*inch])
    t.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('ALIGN', (1,0), (1,-1), 'RIGHT'),
    ]))
    elements.append(t)
    
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Total Orders:</b> {session.total_orders}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Sales:</b> Rs. {session.total_sales:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"<b>Notes:</b> {session.notes or 'None'}", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=shift_report_{session_id}.pdf"}
    )


@router.get("/export/master/excel")
async def export_master_excel(
    start_date: str,
    end_date: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """
    Generate a master Excel report with separate sheets for each date in the range.
    Each date sheet contains: Sales, Inventory Tracking, Item Tracking, Transaction Log
    """
    from datetime import datetime, timedelta
    import pandas as pd
    from io import BytesIO
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill
    from app.models.inventory import InventoryTransaction, BatchProduction, BillOfMaterials
    
    try:
        start_dt = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    if start_dt > end_dt:
        raise HTTPException(status_code=400, detail="Start date must be before or equal to end date")
    
    branch_id = get_branch_filter(current_user)
    metadata = get_branch_metadata(current_user, db) or {}
    
    
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Iterate through each date in the range
    current_date = start_dt
    while current_date <= end_dt:
        date_str = current_date.strftime('%Y-%m-%d')
        sheet_name = date_str  # Use full date as sheet name
        
        # Get sales data for this date
        from datetime import time
        day_start = datetime.combine(current_date, time.min)
        day_end = datetime.combine(current_date, time.max)
        
        # ===== 1. SALES DATA =====
        orders_query = db.query(Order).options(
            joinedload(Order.items).joinedload(OrderItem.menu_item)
        ).filter(
            Order.created_at.between(day_start, day_end),
            Order.status != 'Cancelled'
        )
        orders_query = apply_branch_filter_order(orders_query, branch_id)
        orders = orders_query.all()
        
        sales_data = []
        for o in orders:
            sales_data.append({
                "Order #": o.order_number,
                "Type": o.order_type or "-",
                "Gross": o.total_amount,
                "Discount": o.discount,
                "Net": o.net_amount,
                "Paid": o.paid_amount,
                "Credit": o.credit_amount,
                "Payment": o.payment_type or "-",
                "Items": ", ".join([f"{i.menu_item.name} x{i.quantity}" for i in o.items if i.menu_item]),
                "Time": o.created_at.strftime('%H:%M') if o.created_at else "-"
            })
        
        # ===== 2. INVENTORY DATA =====
        # Filter products created up to this day
        products = db.query(Product).options(joinedload(Product.unit)).filter(
            Product.created_at <= day_end
        ).all()
        
        txns_query = db.query(InventoryTransaction).filter(
            InventoryTransaction.created_at <= day_end
        ).order_by(InventoryTransaction.created_at.asc())
        all_txns = txns_query.all()
        
        summary_stats = {p.id: {
            "opening": 0.0,
            "added": 0.0,
            "produced": 0.0,
            "consumed": 0.0,
            "sold": 0.0,
            "adjusted": 0.0,
            "last_txn_date": "-"
        } for p in products}
        
        for txn in all_txns:
            pid = txn.product_id
            if pid not in summary_stats:
                continue
            
            qty = abs(txn.quantity or 0.0)
            t_type = txn.transaction_type
            
            is_today = day_start <= txn.created_at <= day_end
            
            if not is_today:
                # Contribute to opening stock
                if t_type in ['Purchase_IN', 'Production_IN', 'IN', 'Add']:
                    summary_stats[pid]["opening"] += qty
                elif t_type in ['Production_OUT', 'Sale_OUT', 'OUT', 'Remove']:
                    summary_stats[pid]["opening"] -= qty
                elif t_type in ['Adjustment', 'Count']:
                    summary_stats[pid]["opening"] += txn.quantity
                continue

            # Activity specifically for today
            if t_type in ['Purchase_IN', 'IN', 'Add']:
                summary_stats[pid]["added"] += qty
            elif t_type == 'Production_IN':
                summary_stats[pid]["produced"] += qty
            elif t_type == 'Production_OUT':
                summary_stats[pid]["consumed"] += qty
            elif t_type == 'Sale_OUT':
                summary_stats[pid]["sold"] += qty
            elif t_type in ['Adjustment', 'Count']:
                summary_stats[pid]["adjusted"] += txn.quantity
            
            summary_stats[pid]["last_txn_date"] = txn.created_at.strftime('%Y-%m-%d')
        
        inventory_data = []
        for p in products:
            stats = summary_stats[p.id]
            movements = (stats["added"] + stats["produced"] + stats["adjusted"]) - (stats["consumed"] + stats["sold"])
            closing = stats["opening"] + movements
            
            inventory_data.append({
                "Product": p.name,
                "Category": p.category or "-",
                "Unit": p.unit.abbreviation if p.unit else "-",
                "Opening": round(stats["opening"], 2),
                "Added": round(stats["added"], 2),
                "Produced": round(stats["produced"], 2),
                "Consumed": round(stats["consumed"], 2),
                "Sold": round(stats["sold"], 2),
                "Adjusted": round(stats["adjusted"], 2),
                "Closing": round(closing, 2),
                "Last Txn": stats["last_txn_date"]
            })
        
        # ===== 3. ITEM TRACKING DATA =====
        all_batches = db.query(BatchProduction).options(
            joinedload(BatchProduction.bom).joinedload(BillOfMaterials.menu_items)
        ).filter(BatchProduction.created_at <= day_end).all()
        
        menu_item_stats = {}
        for batch in all_batches:
            if not batch.bom:
                continue
            
            menu_items = batch.bom.menu_items if batch.bom.menu_items else []
            
            for mi in menu_items:
                if mi.id not in menu_item_stats:
                    sold = db.query(func.sum(OrderItem.quantity)).join(Order).filter(
                        OrderItem.menu_item_id == mi.id,
                        Order.status != 'Cancelled',
                        Order.created_at <= day_end
                    ).scalar()
                    total_sold = float(sold) if sold else 0.0
                    
                    total_prod_for_item = 0.0
                    for b in all_batches:
                        if b.bom and any(m.id == mi.id for m in b.bom.menu_items):
                            total_prod_for_item += b.bom.output_quantity * b.quantity
                    
                    menu_item_stats[mi.id] = {
                        "name": mi.name,
                        "produced": total_prod_for_item,
                        "sold": total_sold,
                        "remaining": total_prod_for_item - total_sold
                    }
        
        item_tracking_data = []
        for mi_id, mi_stats in menu_item_stats.items():
            item_tracking_data.append({
                "Menu Item": mi_stats["name"],
                "Produced": round(mi_stats["produced"], 2),
                "Sold": round(mi_stats["sold"], 2),
                "Remaining": round(mi_stats["remaining"], 2)
            })
        
        # ===== 4. PURCHASE DATA =====
        from app.models.purchase import PurchaseBill, PurchaseBillItem
        purchases_query = db.query(PurchaseBill).options(
            joinedload(PurchaseBill.supplier),
            joinedload(PurchaseBill.items).joinedload(PurchaseBillItem.product)
        ).filter(
            PurchaseBill.order_date.between(day_start, day_end)
        )
        if branch_id:
            purchases_query = purchases_query.filter(PurchaseBill.branch_id == branch_id)
        purchases = purchases_query.all()
        
        purchase_data = []
        for p in purchases:
            items_str = ", ".join([f"{i.product.name if i.product else 'Unknown'} ({i.quantity} {i.unit.abbreviation if i.unit else ''})" for i in p.items])
            paid = p.total_amount if getattr(p, 'status', '').lower() == 'paid' else 0
            due = p.total_amount - paid
            purchase_data.append({
                "Bill #": p.bill_number,
                "Supplier": p.supplier.name if p.supplier else "-",
                "Items": items_str,
                "Total": p.total_amount,
                "Paid": paid,
                "Due": due,
                "Date": p.order_date.strftime('%Y-%m-%d') if p.order_date else "-"
            })
        
        # ===== 5. POS SESSION DATA =====
        from app.models.pos_session import POSSession
        session_query = db.query(POSSession).options(joinedload(POSSession.user)).filter(
            (POSSession.start_time.between(day_start, day_end)) | 
            (POSSession.end_time.between(day_start, day_end))
        )
        if branch_id:
            session_query = session_query.filter(POSSession.branch_id == branch_id)
        sessions_list = session_query.all()
        
        session_data = []
        for s in sessions_list:
            session_data.append({
                "Staff": s.user.full_name if s.user else "System",
                "Status": s.status,
                "Start": s.start_time.strftime('%H:%M') if s.start_time else "-",
                "End": s.end_time.strftime('%H:%M') if s.end_time else "Open",
                "Opening Cash": s.opening_cash,
                "Actual Reported": s.actual_cash,
                "Expected Cash": s.expected_cash,
                "Difference": s.actual_cash - s.expected_cash,
                "Total Sales": s.total_sales
            })
        
        # ===== CREATE SHEET AND WRITE DATA =====
        ws = wb.create_sheet(title=sheet_name)
        
        # Create DataFrames
        df_sales = pd.DataFrame(sales_data) if sales_data else pd.DataFrame([{"Order #": "No sales"}])
        df_inventory = pd.DataFrame(inventory_data) if inventory_data else pd.DataFrame([{"Product": "No inventory"}])
        df_items = pd.DataFrame(item_tracking_data) if item_tracking_data else pd.DataFrame([{"Menu Item": "No items"}])
        df_purchases = pd.DataFrame(purchase_data) if purchase_data else pd.DataFrame([{"Bill #": "No purchases"}])
        df_sessions = pd.DataFrame(session_data) if session_data else pd.DataFrame([{"Staff": "No sessions"}])
        
        # Write sections horizontally with gaps
        col_offset = 1
        
        # Section 1: Sales
        ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+len(df_sales.columns)-1)
        header_cell = ws.cell(row=1, column=col_offset)
        header_cell.value = "DAILY SALES"
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
        
        for r_idx, row in enumerate(dataframe_to_rows(df_sales, index=False, header=True), 2):
            for c_idx, value in enumerate(row, col_offset):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
        
        col_offset += len(df_sales.columns) + 3
        
        # Section 2: Inventory
        ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+len(df_inventory.columns)-1)
        header_cell = ws.cell(row=1, column=col_offset)
        header_cell.value = "INVENTORY TRACKING"
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        
        for r_idx, row in enumerate(dataframe_to_rows(df_inventory, index=False, header=True), 2):
            for c_idx, value in enumerate(row, col_offset):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
        
        col_offset += len(df_inventory.columns) + 3
        
        # Section 3: Items
        ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+len(df_items.columns)-1)
        header_cell = ws.cell(row=1, column=col_offset)
        header_cell.value = "ITEM TRACKING"
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        
        for r_idx, row in enumerate(dataframe_to_rows(df_items, index=False, header=True), 2):
            for c_idx, value in enumerate(row, col_offset):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
        
        col_offset += len(df_items.columns) + 3
        
        # Section 4: Purchases
        ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+len(df_purchases.columns)-1)
        header_cell = ws.cell(row=1, column=col_offset)
        header_cell.value = "PURCHASE REPORT"
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid')
        
        for r_idx, row in enumerate(dataframe_to_rows(df_purchases, index=False, header=True), 2):
            for c_idx, value in enumerate(row, col_offset):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')

        col_offset += len(df_purchases.columns) + 3
        
        # Section 5: Sessions
        ws.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset+len(df_sessions.columns)-1)
        header_cell = ws.cell(row=1, column=col_offset)
        header_cell.value = "POS SESSIONS"
        header_cell.font = Font(bold=True, size=12)
        header_cell.alignment = Alignment(horizontal='center')
        header_cell.fill = PatternFill(start_color='22c55e', end_color='22c55e', fill_type='solid')
        
        for r_idx, row in enumerate(dataframe_to_rows(df_sessions, index=False, header=True), 2):
            for c_idx, value in enumerate(row, col_offset):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
        
        # Move to next date
        current_date += timedelta(days=1)
    
    # Save workbook to buffer
    buffer = BytesIO()
    wb.save(buffer)
    
    
    buffer.seek(0)
    
    filename = f"Master_Report_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
